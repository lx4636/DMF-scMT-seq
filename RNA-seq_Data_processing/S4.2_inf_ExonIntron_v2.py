# !/usr/bin/env python
# -*- coding=utf-8 -*-

import os
import re
import openpyxl
import argparse

parser = argparse.ArgumentParser(description='The ExonIntron information of each sample was integrated: \
                                              python2 inf_ExonIntron_v2.py \
                                              -o output')

parser.add_argument('-o', '--output', type=str, required=True, help='output_file_name')

args = parser.parse_args()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws['A1'] = '% of bases mapped to exon'
ws['B1'] = '% of bases mapped to intron'
ws['C1'] = '% of bases mapped to gene'
ws['D1'] = 'Exon bases'
ws['E1'] = 'Gene bases'
ws['F1'] = 'Total mapped bases'
ws['G1'] = 'Sample name'
count = 2
os.system("ls *_ExonIntron.txt > list_ExonIntron_tmp.txt")
file_list = open('list_ExonIntron_tmp.txt', 'r')
for i in file_list:
    m = re.sub('\n', '', i)
    file_ExonIntron = open(m, 'r')
    for i2 in file_ExonIntron:
        m2 = re.sub('\n', '', i2)
        n = m2.split(': ')
        if n[0] == 'Sample name':
            Sample_name = n[1]
        elif n[0] == 'Total mapped bases':
            Total_mapped_bases = n[1]
        elif n[0] == 'Exon bases':
            Exon_bases = n[1]
        elif n[0] == '% of bases mapped to exon':
            p_of_bases_mapped_to_exon = n[1]
        elif n[0] == 'Gene bases':
            Gene_bases = n[1]
        elif n[0] == '% of bases mapped to gene':
            p_of_bases_mapped_to_gene = n[1]
        elif n[0] == '% of bases mapped to intron':
            p_of_bases_mapped_to_intron = n[1]
    ws['A'+str(count)] = p_of_bases_mapped_to_exon
    ws['B'+str(count)] = p_of_bases_mapped_to_intron
    ws['C'+str(count)] = p_of_bases_mapped_to_gene
    ws['D'+str(count)] = Exon_bases
    ws['E'+str(count)] = Gene_bases
    ws['F'+str(count)] = Total_mapped_bases
    ws['G'+str(count)] = Sample_name
    count = count+1
file_list.close()
os.system("rm list_ExonIntron_tmp.txt")
wb.save(args.output+'_ExonIntron_MAP.xlsx')
