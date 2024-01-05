#!/usr/bin/env python
#coding=utf-8

import os
import re
import numpy as np
import codecs
import openpyxl
import argparse

parser = argparse.ArgumentParser(description='At the same sequencing depth, the number of genes detected was calculated \
                                              python2 DepthGeneReadCount_v3.py \
                                              -s list_Aligned.out.sam \
                                              -g url_to_gtf \
                                              -d deptha \
                                              -o output')

parser.add_argument('-s', '--sam', type=str, default=0, help='a list of Aligned.out.sam')
parser.add_argument('-g', '--gtf', type=str, default='/home/songjia/reference/hg19/gencode.v35lift37.annotation.gtf', help='gtf for htseq-count')
parser.add_argument('-d', '--depth', type=str, required=True, help='depth')
parser.add_argument('-o', '--output', type=str, required=True, help='output_depth_GeneReadCount.xls')

args = parser.parse_args()

def DepthGene(sam, fq, d):
    file_fq = open(fq, 'r')
    cont_fq = file_fq.readlines()
    nn = 0
    nn_max = len(cont_fq)
    pattern = r'@(.*?)\s'
    list_id = []
    while (nn<nn_max):
        if nn%4==0:
            idid = re.match(pattern, cont_fq[nn])
            list_id.append(idid.group(1))
        nn = nn+1
    list_ID = np.random.choice(list_id, d, replace = False)
    dict0 = {}
    for i in list_id:
        dict0[i] = 0
    for i in list_ID:
        dict0[i] = 1
    file_fq.close()

    file_align = open(sam, 'r')
    prefix = re.sub('Aligned.out.sam','',sam)
    file_align_id = codecs.open(prefix+'_'+str(d)+'DepthGene.out.sam','w')
    for line in file_align:
        if line.startswith('@'):
            file_align_id.write(line)
        else:
            ll = line.split('\t')
            if dict0[ll[0]]:
                file_align_id.write(line)
    file_align.close()
    file_align_id.close()

def GeneReadCount(htseq_out):
    file = open(htseq_out,'r')
    prefix = re.sub('_counts.txt','',htseq_out)
    file_count = codecs.open(prefix+'_GeneReadCount.txt','w')
    count=0
    for line in file:
        if ('processed' in line) or ('__' in line):
            continue
        else:
            line2 = re.sub('\n','',line)
            l = line2.split('\t')
            if (l[1] != '0'):
                count=count+1
                file_count.write(l[0]+'\t'+l[1]+'\n')
    file_count.write('Gene Count: '+str(count))
    file.close()
    file_count.close()

if args.sam:
    list_sam=[]
    file=open(args.sam,'r')
    for i in file:
        sam_name=re.sub('\n','',i)
        list_sam.append(sam_name)
else:
    os.system("ls *Aligned.out.sam > list_Aligned.out.sam")
    list_sam=[]
    file=open('list_Aligned.out.sam','r')
    for i in file:
        sam_url=re.sub('\n','',i)
        list_sam.append(sam_url)
    os.system("rm list_Aligned.out.sam")

d = int(args.depth)
gtf = args.gtf
output = args.output
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
row=3
for i in list_sam:
    sam=i
    fq=re.sub('Aligned.out.sam','_1_repair_1.fq',i)
    url_fq='../'+fq
    DepthGene(sam, url_fq, d)
    prefix = re.sub('Aligned.out.sam','',sam)
    DepthGene_out=prefix+'_'+str(d)+'DepthGene.out.sam'
    
    htseq_out=prefix+'_'+str(d)+'_counts.txt'
    os.system("htseq-count -f sam -r name -s no -a 10 -t gene -i gene_id -m intersection-nonempty %s %s > %s" %(DepthGene_out, gtf, htseq_out))
    result=os.popen("wc -l %s" %(htseq_out))
    res=result.read()
    num=res.split(' ')[0]
    total_gene_num=int(num)-5

    GeneReadCount(htseq_out)
    GeneReadCount_out=prefix+'_'+str(d)+'_GeneReadCount.txt'
    file_GeneReadCount=open(GeneReadCount_out,'r')
    cont=file_GeneReadCount.readlines()
    last_line=cont[-1].split(': ')
    gene_count=last_line[1]
    ws['A'+str(row)] = prefix
    ws['B'+str(row)] = gene_count
    print(i+' is finished!')
    row=row+1
    
ws['A1'] = 'depth_reads '+str(d)
ws['B1'] = 'total_gene_num '+str(total_gene_num)
ws['A2'] = 'sample_name'
ws['B2'] = 'gene_count'
wb.save(output+'_'+str(d)+'_GeneReadCount.xlsx')

